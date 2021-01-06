from openpyxl import load_workbook
import xThermoInterface as xt
import optimizer as opt

# Create a new instance of the thermodynamic property engine (Xiaodong's wrapper for xThermo.dll).
thermo = xt.xThermoInterface()

# Select the CPA equation of state.
thermo.ChooseAModel(1)

# Specify the number of components.
nc = 1
thermo.NoPureComp(nc)

# Load saturation properties workbook in read only mode with cell values (not formulas).
wb = load_workbook('saturation_properties.xlsx', read_only=False, data_only=True)

# Prepare a worksheet to store results
results_sheet = wb['CPA_Results']

# Compounds and critical constants.
crit_props_sheet = wb['Critical']
compounds = []
crit_props = {}
for row in crit_props_sheet.iter_rows(min_row=2, min_col=1, max_col=8):
    compounds.append(row[0].value)
    # Append critical properties list [Tc [=] K, Pc [=] bar, omega [=] dimensionless].
    crit_props[row[0].value] = [row[1].value, row[7].value, row[5].value]

# Cubic Plus Association (CPA) constants for initial guesses.
cpa_consts_sheet = wb['CPA_Reoptimized']
cpa_consts = {}
for row in cpa_consts_sheet.iter_rows(min_row=3, min_col=1, max_col=19):
    # A set of CPA constants for each pure component is appended to the cpa_consts dictionary.
    #
    # Pure component sets are  formatted in a utilitarian and rather lazy way. A pure component is really best handled
    # as its own class. Rebuilding this code in an object oriented way will abandon the current implementation in
    # favor of a pure component class.
    #
    # Dictionary format:
    #     {const_identifier: [comp_string, physical_parm_list, assoc_parm_list, assoc_scheme_int]}
    #     compound_string = compound name string
    #     physical_parm_list = [b*1000, gamma, c1]
    #     assoc_parm_list = [beta*1000, epsilon]
    #     assoc_vol = beta*1000
    #     assoc_energy = epsilon
    #     assoc_scheme_int = 11 (2B), 12 (3B), 22 (4C), 100 (1A)
    cpa_consts[row[0].value] = [row[1].value,
                                [row[14].value, row[15].value, row[16].value],
                                [row[17].value, row[18].value],
                                row[17].value,
                                row[18].value,
                                row[7].value]


def cpa_sat(t, b, gamma, c1, beta, epsilon, assoc_scheme, prop_type):
    """Calculate the saturation pressure and liquid density.
    Args:
        t (float): Saturation temperature.
        b (float): Pure component CPA parameter.
        gamma (float): Pure component CPA parameter.
        c1 (float): Pure component CPA parameter.
        beta (float): Pure component CPA parameter.
        epsilon (float): Pure component CPA parameter.
        assoc_scheme (int): CPA association scheme.
        prop_type (string): Property to be calculated.
    Returns:
        prop_value: Returns saturation pressure (Pa) or liquid density (kmol/m^3).
    """
    # Pass physical and association parameters to the thermodynamic property engine.
    thermo.CPAParams(1, b, gamma, c1)
    thermo.AssocParams(1, assoc_scheme, beta, epsilon)
    # Fix model inputs passed to the thermodynamic property engine (required before property evaluation).
    thermo.Setup_Thermo()
    # Calculate the saturation pressure.
    psat, ln_k, ierr = thermo.PBubble(t, [1.0])
    # Calculate the liquid compressibility factor at saturated conditions.
    z_l, _ = thermo.FugacityCoeff(t, psat, [1.0], iph=1, job=0)
    # Universal gas constant, J/mol.K
    r = 8.3145
    # Calculate molar liquid density, mol/m^3
    rho_l_molar = (100000.0 * psat) / (z_l * r * t)
    if prop_type == 'Saturation Pressure':
        return psat * 100000.0  # Saturation pressure [=] Pa
    elif prop_type == 'Saturation Liquid Density':
        return rho_l_molar / 1000.0  # Saturation liquid density [=] kmol/m^3
    else:
        raise ValueError('prop_type not valid')


for const_id in cpa_consts.keys():
    # Pure component parameters for the CPA EOS.
    comp = cpa_consts[const_id][0]
    cpa_phys = cpa_consts[const_id][1]
    cpa_assoc = cpa_consts[const_id][2]
    cpa_assoc_vol = cpa_consts[const_id][3]
    cpa_assoc_energy = cpa_consts[const_id][4]
    cpa_scheme = cpa_consts[const_id][5]
    cpa_parms = cpa_phys + cpa_assoc

    # Critical properties for the compound.
    tc = crit_props[comp][0]
    pc = crit_props[comp][1]
    omega = crit_props[comp][2]

    # Pass critical properties to the thermodynamic property engine.
    thermo.CritProps(1, tc, pc, omega)

    # Retrieve saturation pressure and liquid density data.
    sheet = wb[comp]
    type = []
    unit = []
    source = []
    temp = []
    value = []
    unc = []
    unc_frac = []
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=7):
        # Append data to lists only if the 'value' is not None or 0.
        # DIPPR correlation pseudo-data is specified here.
        # Other data sources can be specified --> ['NIST Web Thermo', 'NIST REFPROP', 'EXPERIMENTAL']
        if row[2].value == 'DIPPR':
            type.append(row[0].value)
            unit.append(row[1].value)
            source.append(row[2].value)
            temp.append(row[3].value)
            value.append(row[4].value)
            unc.append(row[5].value)
            unc_frac.append(row[6].value)

    # Weighting parameter for least squares parameter fitting.
    weight = [1.0 / (v ** 2.0) for v in value]

    # Define a list of temperatures, values, and property types for parameter regression.
    temp_type_value = list(zip(temp, type, value))

    # Define the association scheme.  Site order is glue, positive, negative (ex. 022 = 4C, 011 = 2B)
    assoc_scheme = [cpa_scheme] * len(type)

    # Define an args list that can be passed to the least squares optimizer.
    cpa_sat_args = list(zip(assoc_scheme, type))

    # The Nelder-Mead simplex algorithm searches for an optimum in the neighborhood of the initial point.  The initial
    # guess for all compounds is the result of prior optimizations, so it is already very close to the minimum.
    nm_theta = opt.nelder_mead(cpa_parms,
                               opt.least_squares_objective_function,
                               args=(cpa_sat, temp, value, weight, None, cpa_sat_args),
                               max_iter=2500,
                               initial_size=0.05)

    # Test regressed CPA parameters.
    abs_dev_psat = []
    abs_dev_lden = []
    for i in temp_type_value:
        prop_temp = i[0]
        prop_type = i[1]
        prop_value = i[2]
        prop_calc_value = cpa_sat(prop_temp, *nm_theta, cpa_scheme, prop_type)
        if prop_type == 'Saturation Pressure':
            abs_dev_psat.append(abs((prop_calc_value - prop_value)/prop_value))
        if prop_type == 'Saturation Liquid Density':
            abs_dev_lden.append(abs((prop_calc_value - prop_value) / prop_value))

    pct_abs_dev_psat = 100.0*sum(abs_dev_psat)/len(abs_dev_psat)
    pct_abs_dev_lden = 100.0*sum(abs_dev_lden)/len(abs_dev_lden)

    # Save results in spreadsheet.
    results_sheet.append([const_id, 'DIPPR', *nm_theta, pct_abs_dev_lden, pct_abs_dev_psat])
    wb.save('saturation_properties.xlsx')
    print(const_id)
